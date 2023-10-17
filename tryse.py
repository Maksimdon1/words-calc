from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
























#aa = 'абвгдеёжзийклмнопрстуфхцчшщъыьэюя'
enddict = {}
data = ['а', 'б', 'в', 'г', 'д', 'е', 'ё', 'ж', 'з', 'и', 'й', 'к', 'л', 'м', 'н', 'о', 'п', 'р', 'с', 'т', 'у', 'ф', 'х', 'ц', 'ч', 'ш', 'щ', 'ъ', 'ы', 'ь', 'э', 'ю', 'я']
thisdict = {'а': 0, 'б': 0, 'в': 0, 'г': 0, 'д': 0, 'е': 0, 'ё': 0, 'ж': 0, 'з': 0, 'и': 0, 'й': 0, 'к': 0, 'л': 0, 'м': 0, 'н': 0, 'о': 0, 'п': 0, 'р': 0, 'с': 0, 'т': 0, 'у': 0, 'ф': 0, 'х': 0, 'ц': 0, 'ч': 0, 'ш': 0, 'щ': 0, 'ъ': 0, 'ы': 0, 'ь': 0, 'э': 0, 'ю': 0, 'я': 0}
text = """
Три девицы под окном
Пряли поздно вечерком.
«Кабы я была царица, —
Говорит одна девица, —
То на весь крещеный мир
Приготовила б я пир».
«Кабы я была царица, —
Говорит ее сестрица, —
То на весь бы мир одна
Наткала я полотна».
«Кабы я была царица, —
Третья молвила сестрица, —
Я б для батюшки-царя
Родила богатыря».
Только вымолвить успела,
Дверь тихонько заскрыпела,
И в светлицу входит царь,
Стороны той государь.
Во всё время разговора
Он стоял позадь забора;
Речь последней по всему
Полюбилася ему.
«Здравствуй, красная девица, —
Говорит он, — будь царица
И роди богатыря
Мне к исходу сентября.
Вы ж, голубушки-сестрицы,
Выбирайтесь из светлицы,
Поезжайте вслед за мной,
Вслед за мной и за сестрой:
Будь одна из вас ткачиха,
А другая повариха».
В сени вышел царь-отец.
Все пустились во дворец.
Царь недолго собирался:
В тот же вечер обвенчался.
Царь Салтан за пир честной
Сел с царицей молодой;
А потом честные гости
На кровать слоновой кости
Положили молодых
И оставили одних.
В кухне злится повариха,
Плачет у станка ткачиха,
И завидуют оне
Государевой жене.
А царица молодая,
Дела вдаль не отлагая,
С первой ночи понесла.
В те поры война была.
Царь Салтан, с женой простяся,
На добра-коня садяся,
Ей наказывал себя
Поберечь, его любя.
Между тем, как он далёко
Бьется долго и жестоко,
Наступает срок родин;
Сына бог им дал в аршин,
И царица над ребенком
Как орлица над орленком;
Шлет с письмом она гонца,
Чтоб обрадовать отца.
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Извести ее хотят,
Перенять гонца велят;
Сами шлют гонца другого
Вот с чем от слова до слова:
«Родила царица в ночь
Не то сына, не то дочь;
Не мышонка, не лягушку,
А неведому зверюшку».
Как услышал царь-отец,
Что донес ему гонец,
В гневе начал он чудесить
И гонца хотел повесить;
Но, смягчившись на сей раз,
Дал гонцу такой приказ:
«Ждать царева возвращенья
Для законного решенья».
Едет с грамотой гонец,
И приехал наконец.
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Обобрать его велят;
Допьяна гонца поят
И в суму его пустую
Суют грамоту другую —
И привез гонец хмельной
В тот же день приказ такой:
«Царь велит своим боярам,
Времени не тратя даром,
И царицу и приплод
Тайно бросить в бездну вод».
Делать нечего: бояре,
Потужив о государе
И царице молодой,
В спальню к ней пришли толпой.
Объявили царску волю —
Ей и сыну злую долю,
Прочитали вслух указ,
И царицу в тот же час
В бочку с сыном посадили,
Засмолили, покатили
И пустили в Окиян —
Так велел-де царь Салтан.
В синем небе звезды блещут,
В синем море волны хлещут;
Туча по небу идет,
Бочка по морю плывет.
Словно горькая вдовица,
Плачет, бьется в ней царица;
И растет ребенок там
Не по дням, а по часам.
День прошел, царица вопит…
А дитя волну торопит:
«Ты, волна моя, волна!
Ты гульлива и вольна;
Плещешь ты, куда захочешь,
Ты морские камни точишь,
Топишь берег ты земли,
Подымаешь корабли —
Не губи ты нашу душу:
Выплесни ты нас на сушу!»
И послушалась волна:
Тут же на берег она
Бочку вынесла легонько
И отхлынула тихонько.
Мать с младенцем спасена;
Землю чувствует она.
Но из бочки кто их вынет?
Бог неужто их покинет?
Сын на ножки поднялся,
В дно головкой уперся,
Понатужился немножко:
«Как бы здесь на двор окошко
Нам проделать?» — молвил он,
Вышиб дно и вышел вон.
Мать и сын теперь на воле;
Видят холм в широком поле,
Море синее кругом,
Дуб зеленый над холмом.
Сын подумал: добрый ужин
Был бы нам, однако, нужен.
Ломит он у дуба сук
И в тугой сгибает лук,
Со креста снурок шелковый
Натянул на лук дубовый,
Тонку тросточку сломил,
Стрелкой легкой завострил
И пошел на край долины
У моря искать дичины.
К морю лишь подходит он,
Вот и слышит будто стон…
Видно на море не тихо;
Смотрит — видит дело лихо:
Бьется лебедь средь зыбей,
Коршун носится над ней;
Та бедняжка так и плещет,
Воду вкруг мутит и хлещет…
Тот уж когти распустил,
Клёв кровавый навострил…
Но как раз стрела запела,
В шею коршуна задела —
Коршун в море кровь пролил,
Лук царевич опустил;
Смотрит: коршун в море тонет
И не птичьим криком стонет,
Лебедь около плывет,
Злого коршуна клюет,
Гибель близкую торопит,
Бьет крылом и в море топит —
И царевичу потом
Молвит русским языком:
«Ты, царевич, мой спаситель,
Мой могучий избавитель,
Не тужи, что за меня
Есть не будешь ты три дня,
Что стрела пропала в море;
Это горе — всё не горе.
Отплачу тебе добром,
Сослужу тебе потом:
Ты не лебедь ведь избавил,
Девицу в живых оставил;
Ты не коршуна убил,
Чародея подстрелил.
Ввек тебя я не забуду:
Ты найдешь меня повсюду,
А теперь ты воротись,
Не горюй и спать ложись».
Улетела лебедь-птица,
А царевич и царица,
Целый день проведши так,
Лечь решились на тощак.
Вот открыл царевич очи;
Отрясая грезы ночи
И дивясь, перед собой
Видит город он большой,
Стены с частыми зубцами,
И за белыми стенами
Блещут маковки церквей
И святых монастырей.
Он скорей царицу будит;
Та как ахнет!.. «То ли будет? —
Говорит он, — вижу я:
Лебедь тешится моя».
Мать и сын идут ко граду.
Лишь ступили за ограду,
Оглушительный трезвон
Поднялся со всех сторон:
К ним народ навстречу валит,
Хор церковный бога хвалит;
В колымагах золотых
Пышный двор встречает их;
Все их громко величают
И царевича венчают
Княжей шапкой, и главой
Возглашают над собой;
И среди своей столицы,
С разрешения царицы,
В тот же день стал княжить он
И нарекся: князь Гвидон.
Ветер на море гуляет
И кораблик подгоняет;
Он бежит себе в волнах
На раздутых парусах.
Корабельщики дивятся,
На кораблике толпятся,
На знакомом острову
Чудо видят наяву:
Город новый златоглавый,
Пристань с крепкою заставой;
Пушки с пристани палят,
Кораблю пристать велят.
Пристают к заставе гости;
Князь Гвидон зовет их в гости,
Их он кормит и поит
И ответ держать велит:
«Чем вы, гости, торг ведете
И куда теперь плывете?»
Корабельщики в ответ:
«Мы объехали весь свет,
Торговали соболями,
Чернобурыми лисами;
А теперь нам вышел срок,
Едем прямо на восток,
Мимо острова Буяна,
В царство славного Салтана…»
Князь им вымолвил тогда:
«Добрый путь вам, господа,
По морю по Окияну
К славному царю Салтану;
От меня ему поклон».
Гости в путь, а князь Гвидон
С берега душой печальной
Провожает бег их дальный;
Глядь — поверх текучих вод
Лебедь белая плывет.
«Здравствуй, князь ты мой прекрасный!
Что ты тих, как день ненастный?
Опечалился чему?» —
Говорит она ему.
Князь печально отвечает:
«Грусть-тоска меня съедает,
Одолела молодца:
Видеть я б хотел отца».
Лебедь князю: «Вот в чем горе!
Ну, послушай: хочешь в море
Полететь за кораблем?
Будь же, князь, ты комаром».
И крылами замахала,
Воду с шумом расплескала
И обрызгала его
С головы до ног всего.
Тут он в точку уменьшился,
Комаром оборотился,
Полетел и запищал,
Судно на море догнал,
Потихоньку опустился
На корабль — и в щель забился.
Ветер весело шумит,
Судно весело бежит
Мимо острова Буяна,
К царству славного Салтана,
И желанная страна
Вот уж издали видна.
Вот на берег вышли гости;
Царь Салтан зовет их в гости,
И за ними во дворец
Полетел наш удалец.
Видит: весь сияя в злате,
Царь Салтан сидит в палате
На престоле и в венце
С грустной думой на лице;
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Около царя сидят
И в глаза ему глядят.
Царь Салтан гостей сажает
За свой стол и вопрошает:
«Ой вы, гости-господа,
Долго ль ездили? куда?
Ладно ль за морем, иль худо?
И какое в свете чудо?»
Корабельщики в ответ:
«Мы объехали весь свет;
За морем житье не худо,
В свете ж вот какое чудо:
В море остров был крутой,
Не привальный, не жилой;
Он лежал пустой равниной;
Рос на нем дубок единый;
А теперь стоит на нем
Новый город со дворцом,
С златоглавыми церквами,
С теремами и садами,
А сидит в нем князь Гвидон;
Он прислал тебе поклон».
Царь Салтан дивится чуду;
Молвит он: «Коль жив я буду,
Чудный остров навещу,
У Гвидона погощу».
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Не хотят его пустить
Чудный остров навестить.
«Уж диковинка, ну право, —
Подмигнув другим лукаво,
Повариха говорит, —
Город у моря стоит!
Знайте, вот что не безделка:
Ель в лесу, под елью белка,
Белка песенки поет
И орешки всё грызет,
А орешки не простые,
Всё скорлупки золотые,
Ядра — чистый изумруд;
Вот что чудом-то зовут».
Чуду царь Салтан дивится,
А комар-то злится, злится —
И впился комар как раз
Тетке прямо в правый глаз.
Повариха побледнела,
Обмерла и окривела.
Слуги, сватья и сестра
С криком ловят комара.
«Распроклятая ты мошка!
Мы тебя!..» А он в окошко,
Да спокойно в свой удел
Через море полетел.
Снова князь у моря ходит,
С синя моря глаз не сводит;
Глядь — поверх текучих вод
Лебедь белая плывет.
«Здравствуй, князь ты мой прекрасный!
Что ж ты тих, как день ненастный?
Опечалился чему?« —
Говорит она ему.
Князь Гвидон ей отвечает:
«Грусть-тоска меня съедает;
Чудо чудное завесть
Мне б хотелось. Где-то есть
Ель в лесу, под елью белка;
Диво, право, не безделка —
Белка песенки поет,
Да орешки всё грызет,
А орешки не простые,
Всё скорлупки золотые,
Ядра — чистый изумруд;
Но, быть может, люди врут».
Князю лебедь отвечает:
«Свет о белке правду бает;
Это чудо знаю я;
Полно, князь, душа моя,
Не печалься; рада службу
Оказать тебе я в дружбу».
С ободренною душой
Князь пошел себе домой;
Лишь ступил на двор широкий —
Что ж? под елкою высокой,
Видит, белочка при всех
Золотой грызет орех,
Изумрудец вынимает,
А скорлупку собирает,
Кучки равные кладет
И с присвисточкой поет
При честном при всем народе:
Во саду ли, в огороде.
Изумился князь Гвидон.
«Ну, спасибо, — молвил он, —
Ай да лебедь — дай ей боже,
Что и мне, веселье то же».
Князь для белочки потом
Выстроил хрустальный дом,
Караул к нему приставил
И притом дьяка заставил
Строгий счет орехам весть.
Князю прибыль, белке честь.
Ветер по морю гуляет
И кораблик подгоняет;
Он бежит себе в волнах
На поднятых парусах
Мимо острова крутого,
Мимо города большого:
Пушки с пристани палят,
Кораблю пристать велят.
Пристают к заставе гости;
Князь Гвидон зовет их в гости,
Их и кормит и поит
И ответ держать велит:
«Чем вы, гости, торг ведете
И куда теперь плывете?»
Корабельщики в ответ:
«Мы объехали весь свет,
Торговали мы конями,
Всё донскими жеребцами,
А теперь нам вышел срок —
И лежит нам путь далек:
Мимо острова Буяна,
В царство славного Салтана…»
Говорит им князь тогда:
«Добрый путь вам, господа,
По морю по Окияну
К славному царю Салтану;
Да скажите: князь Гвидон
Шлет царю-де свой поклон».
Гости князю поклонились,
Вышли вон и в путь пустились.
К морю князь — а лебедь там
Уж гуляет по волнам.
Молит князь: душа-де просит,
Так и тянет и уносит…
Вот опять она его
Вмиг обрызгала всего:
В муху князь оборотился,
Полетел и опустился
Между моря и небес
На корабль — и в щель залез.
Ветер весело шумит,
Судно весело бежит
Мимо острова Буяна,
В царство славного Салтана —
И желанная страна
Вот уж издали видна;
Вот на берег вышли гости;
Царь Салтан зовет их в гости,
И за ними во дворец
Полетел наш удалец.
Видит: весь сияя в злате,
Царь Салтан сидит в палате
На престоле и в венце,
С грустной думой на лице.
А ткачиха с Бабарихой
Да с кривою поварихой
Около царя сидят,
Злыми жабами глядят.
Царь Салтан гостей сажает
За свой стол и вопрошает:
«Ой вы, гости-господа,
Долго ль ездили? куда?
Ладно ль за морем, иль худо,
И какое в свете чудо?»
Корабельщики в ответ:
«Мы объехали весь свет;
За морем житье не худо;
В свете ж вот какое чудо:
Остров на море лежит,
Град на острове стоит
С златоглавыми церквами,
С теремами да садами;
Ель растет перед дворцом,
А под ней хрустальный дом;
Белка там живет ручная,
Да затейница какая!
Белка песенки поет,
Да орешки всё грызет,
А орешки не простые,
Всё скорлупки золотые,
Ядра — чистый изумруд;
Слуги белку стерегут,
Служат ей прислугой разной —
И приставлен дьяк приказный
Строгий счет орехам весть;
Отдает ей войско честь;
Из скорлупок льют монету,
Да пускают в ход по свету;
Девки сыплют изумруд
В кладовые, да под спуд;
Все в том острове богаты,
Изоб нет, везде палаты;
А сидит в нем князь Гвидон;
Он прислал тебе поклон».
Царь Салтан дивится чуду.
«Если только жив я буду,
Чудный остров навещу,
У Гвидона погощу».
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Не хотят его пустить
Чудный остров навестить.
Усмехнувшись исподтиха,
Говорит царю ткачиха:
«Что тут дивного? ну, вот!
Белка камушки грызет,
Мечет золото и в груды
Загребает изумруды;
Этим нас не удивишь,
Правду ль, нет ли говоришь.
В свете есть иное диво:
Море вздуется бурливо,
Закипит, подымет вой,
Хлынет на берег пустой,
Разольется в шумном беге,
И очутятся на бреге,
В чешуе, как жар горя,
Тридцать три богатыря,
Все красавцы удалые,
Великаны молодые,
Все равны, как на подбор,
С ними дядька Черномор.
Это диво, так уж диво,
Можно молвить справедливо!»
Гости умные молчат,
Спорить с нею не хотят.
Диву царь Салтан дивится,
А Гвидон-то злится, злится…
Зажужжал он и как раз
Тетке сел на левый глаз,
И ткачиха побледнела:
«Ай!» и тут же окривела;
Все кричат: «Лови, лови,
Да дави ее, дави…
Вот ужо! постой немножко,
Погоди…» А князь в окошко,
Да спокойно в свой удел
Через море прилетел.
Князь у синя моря ходит,
С синя моря глаз не сводит;
Глядь — поверх текучих вод
Лебедь белая плывет.
«Здравствуй, князь ты мой прекрасный!
Что ты тих, как день ненастный?
Опечалился чему?» —
Говорит она ему.
Князь Гвидон ей отвечает:
«Грусть-тоска меня съедает —
Диво б дивное хотел
Перенесть я в мой удел».
«А какое ж это диво?»
— Где-то вздуется бурливо
Окиян, подымет вой,
Хлынет на берег пустой,
Расплеснется в шумном беге,
И очутятся на бреге,
В чешуе, как жар горя,
Тридцать три богатыря,
Все красавцы молодые,
Великаны удалые,
Все равны, как на подбор,
С ними дядька Черномор.
Князю лебедь отвечает:
«Вот что, князь, тебя смущает?
Не тужи, душа моя,
Это чудо знаю я.
Эти витязи морские
Мне ведь братья все родные.
Не печалься же, ступай,
В гости братцев поджидай».
Князь пошел, забывши горе,
Сел на башню, и на море
Стал глядеть он; море вдруг
Всколыхалося вокруг,
Расплескалось в шумном беге
И оставило на бреге
Тридцать три богатыря;
В чешуе, как жар горя,
Идут витязи четами,
И, блистая сединами,
Дядька впереди идет
И ко граду их ведет.
С башни князь Гвидон сбегает,
Дорогих гостей встречает;
Второпях народ бежит;
Дядька князю говорит:
«Лебедь нас к тебе послала
И наказом наказала
Славный город твой хранить
И дозором обходить.
Мы отныне ежеденно
Вместе будем непременно
У высоких стен твоих
Выходить из вод морских,
Так увидимся мы вскоре,
А теперь пора нам в море;
Тяжек воздух нам земли».
Все потом домой ушли.
Ветер по морю гуляет
И кораблик подгоняет;
Он бежит себе в волнах
На поднятых парусах
Мимо острова крутого,
Мимо города большого;
Пушки с пристани палят,
Кораблю пристать велят.
Пристают к заставе гости.
Князь Гвидон зовет их в гости,
Их и кормит и поит
И ответ держать велит:
«Чем вы, гости, торг ведете?
И куда теперь плывете?»
Корабельщики в ответ:
«Мы объехали весь свет;
Торговали мы булатом,
Чистым серебром и златом,
И теперь нам вышел срок;
А лежит нам путь далек,
Мимо острова Буяна,
В царство славного Салтана».
Говорит им князь тогда:
«Добрый путь вам, господа,
По морю по Окияну
К славному царю Салтану.
Да скажите ж: князь Гвидон
Шлет-де свой царю поклон».
Гости князю поклонились,
Вышли вон и в путь пустились.
К морю князь, а лебедь там
Уж гуляет по волнам.
Князь опять: душа-де просит…
Так и тянет и уносит…
И опять она его
Вмиг обрызгала всего.
Тут он очень уменьшился,
Шмелем князь оборотился,
Полетел и зажужжал;
Судно на море догнал,
Потихоньку опустился
На корму — и в щель забился.
Ветер весело шумит,
Судно весело бежит
Мимо острова Буяна,
В царство славного Салтана,
И желанная страна
Вот уж издали видна.
Вот на берег вышли гости.
Царь Салтан зовет их в гости,
И за ними во дворец
Полетел наш удалец.
Видит, весь сияя в злате,
Царь Салтан сидит в палате
На престоле и в венце,
С грустной думой на лице.
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Около царя сидят —
Четырьмя все три глядят.
Царь Салтан гостей сажает
За свой стол и вопрошает:
«Ой вы, гости-господа,
Долго ль ездили? куда?
Ладно ль за морем иль худо?
И какое в свете чудо?»
Корабельщики в ответ:
«Мы объехали весь свет;
За морем житье не худо;
В свете ж вот какое чудо:
Остров на море лежит,
Град на острове стоит,
Каждый день идет там диво:
Море вздуется бурливо,
Закипит, подымет вой,
Хлынет на берег пустой,
Расплеснется в скором беге —
И останутся на бреге
Тридцать три богатыря,
В чешуе златой горя,
Все красавцы молодые,
Великаны удалые,
Все равны, как на подбор;
Старый дядька Черномор
С ними из моря выходит
И попарно их выводит,
Чтобы остров тот хранить
И дозором обходить —
И той стражи нет надежней,
Ни храбрее, ни прилежней.
А сидит там князь Гвидон;
Он прислал тебе поклон».
Царь Салтан дивится чуду.
«Коли жив я только буду,
Чудный остров навещу
И у князя погощу».
Повариха и ткачиха
Ни гугу — но Бабариха
Усмехнувшись говорит:
«Кто нас этим удивит?
Люди из моря выходят
И себе дозором бродят!
Правду ль бают, или лгут,
Дива я не вижу тут.
В свете есть такие ль дива?
Вот идет молва правдива:
За морем царевна есть,
Что не можно глаз отвесть:
Днем свет божий затмевает,
Ночью землю освещает,
Месяц под косой блестит,
А во лбу звезда горит.
А сама-то величава,
Выплывает, будто пава;
А как речь-то говорит,
Словно реченька журчит.
Молвить можно справедливо,
Это диво, так уж диво».
Гости умные молчат:
Спорить с бабой не хотят.
Чуду царь Салтан дивится —
А царевич хоть и злится,
Но жалеет он очей
Старой бабушки своей:
Он над ней жужжит, кружится —
Прямо на нос к ней садится,
Нос ужалил богатырь:
На носу вскочил волдырь.
И опять пошла тревога:
«Помогите, ради бога!
Караул! лови, лови,
Да дави его, дави…
Вот ужо! пожди немножко,
Погоди!..» А шмель в окошко,
Да спокойно в свой удел
Через море полетел.
Князь у синя моря ходит,
С синя моря глаз не сводит;
Глядь — поверх текучих вод
Лебедь белая плывет.
«Здравствуй, князь ты мой прекрасный!
Что ж ты тих, как день ненастный?
Опечалился чему?» —
Говорит она ему.
Князь Гвидон ей отвечает:
«Грусть-тоска меня съедает:
Люди женятся; гляжу,
Неженат лишь я хожу».
— А кого же на примете
Ты имеешь? — «Да на свете,
Говорят, царевна есть,
Что не можно глаз отвесть.
Днем свет божий затмевает,
Ночью землю освещает —
Месяц под косой блестит,
А во лбу звезда горит.
А сама-то величава,
Выступает, будто пава;
Сладку речь-то говорит,
Будто реченька журчит.
Только, полно, правда ль это?»
Князь со страхом ждет ответа.
Лебедь белая молчит
И, подумав, говорит:
«Да! такая есть девица.
Но жена не рукавица:
С белой ручки не стряхнешь,
Да за пояс не заткнешь.
Услужу тебе советом —
Слушай: обо всем об этом
Пораздумай ты путем,
Не раскаяться б потом».
Князь пред нею стал божиться,
Что пора ему жениться,
Что об этом обо всем
Передумал он путем;
Что готов душою страстной
За царевною прекрасной
Он пешком идти отсель
Хоть за тридевять земель.
Лебедь тут, вздохнув глубоко,
Молвила: «Зачем далёко?
Знай, близка судьба твоя,
Ведь царевна эта — я».
Тут она, взмахнув крылами,
Полетела над волнами
И на берег с высоты
Опустилася в кусты,
Встрепенулась, отряхнулась
И царевной обернулась:
Месяц под косой блестит,
А во лбу звезда горит;
А сама-то величава,
Выступает, будто пава;
А как речь-то говорит,
Словно реченька журчит.
Князь царевну обнимает,
К белой груди прижимает
И ведет ее скорей
К милой матушки своей.
Князь ей в ноги, умоляя:
«Государыня-родная!
Выбрал я жену себе,
Дочь послушную тебе,
Просим оба разрешенья,
Твоего благословенья:
Ты детей благослови
Жить в совете и любви».
Над главою их покорной
Мать с иконой чудотворной
Слезы льет и говорит:
«Бог вас, дети, наградит».
Князь не долго собирался,
На царевне обвенчался;
Стали жить да поживать,
Да приплода поджидать.
Ветер по морю гуляет
И кораблик подгоняет;
Он бежит себе в волнах
На раздутых парусах
Мимо острова крутого,
Мимо города большого;
Пушки с пристани палят,
Кораблю пристать велят.
Пристают к заставе гости.
Князь Гвидон зовет их в гости,
Он их кормит и поит
И ответ держать велит:
«Чем вы, гости, торг ведете
И куда теперь плывете?»
Корабельщики в ответ:
«Мы объехали весь свет,
Торговали мы недаром
Неуказанным товаром;
А лежит нам путь далек:
Восвояси на восток,
Мимо острова Буяна,
В царство славного Салтана».
Князь им вымолвил тогда:
«Добрый путь вам, господа,
По морю по Окияну
К славному царю Салтану;
Да напомните ему,
Государю своему:
К нам он в гости обещался,
А доселе не собрался —
Шлю ему я свой поклон».
Гости в путь, а князь Гвидон
Дома на сей раз остался
И с женою не расстался.
Ветер весело шумит,
Судно весело бежит
Мимо острова Буяна
К царству славного Салтана,
И знакомая страна
Вот уж издали видна.
Вот на берег вышли гости.
Царь Салтан зовет их в гости.
Гости видят: во дворце
Царь сидит в своем венце,
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Около царя сидят,
Четырьмя все три глядят.
Царь Салтан гостей сажает
За свой стол и вопрошает:
«Ой вы, гости-господа,
Долго ль ездили? куда?
Ладно ль за морем, иль худо?
И какое в свете чудо?»
Корабельщики в ответ:
«Мы объехали весь свет;
За морем житье не худо,
В свете ж вот какое чудо:
Остров на море лежит,
Град на острове стоит,
С златоглавыми церквами,
С теремами и садами;
Ель растет перед дворцом,
А под ней хрустальный дом;
Белка в нем живет ручная,
Да чудесница какая!
Белка песенки поет
Да орешки всё грызет;
А орешки не простые,
Скорлупы-то золотые,
Ядра — чистый изумруд;
Белку холят, берегут.
Там еще другое диво:
Море вздуется бурливо,
Закипит, подымет вой,
Хлынет на берег пустой,
Расплеснется в скором беге,
И очутятся на бреге,
В чешуе, как жар горя,
Тридцать три богатыря,
Все красавцы удалые,
Великаны молодые,
Все равны, как на подбор —
С ними дядька Черномор.
И той стражи нет надежней,
Ни храбрее, ни прилежней.
А у князя женка есть,
Что не можно глаз отвесть:
Днем свет божий затмевает,
Ночью землю освещает;
Месяц под косой блестит,
А во лбу звезда горит.
Князь Гвидон тот город правит,
Всяк его усердно славит;
Он прислал тебе поклон,
Да тебе пеняет он:
К нам-де в гости обещался,
А доселе не собрался».
Тут уж царь не утерпел,
Снарядить он флот велел.
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Не хотят царя пустить
Чудный остров навестить.
Но Салтан им не внимает
И как раз их унимает:
«Что я? царь или дитя? —
Говорит он не шутя: —
Нынче ж еду!» — Тут он топнул,
Вышел вон и дверью хлопнул.
Под окном Гвидон сидит,
Молча на море глядит:
Не шумит оно, не хлещет,
Лишь едва, едва трепещет,
И в лазоревой дали
Показались корабли:
По равнинам Окияна
Едет флот царя Салтана.
Князь Гвидон тогда вскочил,
Громогласно возопил:
«Матушка моя родная!
Ты, княгиня молодая!
Посмотрите вы туда:
Едет батюшка сюда».
Флот уж к острову подходит.
Князь Гвидон трубу наводит:
Царь на палубе стоит
И в трубу на них глядит;
С ним ткачиха с поварихой,
С сватьей бабой Бабарихой;
Удивляются оне
Незнакомой стороне.
Разом пушки запалили;
В колокольнях зазвонили;
К морю сам идет Гвидон;
Там царя встречает он
С поварихой и ткачихой,
С сватьей бабой Бабарихой;
В город он повел царя,
Ничего не говоря.
Все теперь идут в палаты:
У ворот блистают латы,
И стоят в глазах царя
Тридцать три богатыря,
Все красавцы молодые,
Великаны удалые,
Все равны, как на подбор,
С ними дядька Черномор.
Царь ступил на двор широкой:
Там под елкою высокой
Белка песенку поет,
Золотой орех грызет,
Изумрудец вынимает
И в мешочек опускает;
И засеян двор большой
Золотою скорлупой.
Гости дале — торопливо
Смотрят — что ж? княгиня — диво:
Под косой луна блестит,
А во лбу звезда горит;
А сама-то величава,
Выступает, будто пава,
И свекровь свою ведет.
Царь глядит — и узнает…
В нем взыграло ретивое!
«Что я вижу? что такое?
Как!» — и дух в нем занялся…
Царь слезами залился,
Обнимает он царицу,
И сынка, и молодицу,
И садятся все за стол;
И веселый пир пошел.
А ткачиха с поварихой,
С сватьей бабой Бабарихой,
Разбежались по углам;
Их нашли насилу там.
Тут во всем они признались,
Повинились, разрыдались;
Царь для радости такой
Отпустил всех трех домой.
День прошел — царя Салтана
Уложили спать вполпьяна.
Я там был; мед, пиво пил —
И усы лишь обмочил.
"""                  #input('enter text')
words_value = 0
print(len(text))
for i in text.lower() :
    if i in data:
        words_value += 1
        thisdict[i] += 1

print('words value =', words_value)


def create_dict(value, world):
    enddict.update({world: value})

for i in data:
    value = thisdict[i]
    if value != 0:
        create_dict(value=value, world=i)
print(enddict)
print(thisdict)
def bar_chart_with_out_ziro():
    bar_charts = {}
    for i in data:
        try:
            value = enddict[i]
            if value:
                percent = value / words_value
                bar_charts.update({i : percent})

        except:
            pass
    print(bar_charts)
bar_chart_with_out_ziro()

def bar_chart_with_ziro():
    bar_charts = {}
    for i in data:
        try:
            value = thisdict[i]


            percent = value / words_value
            bar_charts.update({i : percent})

        except:
            pass
    print(bar_charts, sep='\n')
    return  bar_charts
values = bar_chart_with_ziro()


















wb = Workbook()

wb.create_sheet(title='Первый лист', index=0)

sheet = wb['Первый лист']
sheet['A1'] = 'Серия 1'
val = 0
for i in data:
    val += 1
    cell = sheet.cell(row = val , column= 1)
    cell.value = values[i]
 

chart = BarChart()
chart.title = 'Заголовок'
data = Reference(sheet,min_col=1, min_row=1 , max_col=1, max_row=33)

chart.add_data(data, titles_from_data=True)

sheet.add_chart(chart, 'C2')
wb.save('example.xlsx')
